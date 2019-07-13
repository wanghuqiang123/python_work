import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

def F(n):
    if n == '基坑开挖':
        return float(0)
    elif n == None:
        return float(0)
    elif n == '/':
        return float(0)
    else:
        return float(n)

def get_source_value(r,str_col):
    return sheet.cell(row=r,column=column_index_from_string(str_col)).value

wb = openpyxl.load_workbook('source_data.xlsx',data_only=True)
wb_result = openpyxl.load_workbook('result1.xlsx')
print('Process...')

sheet = wb.get_sheet_by_name('管网')
sheet_result = wb_result.get_sheet_by_name('Sheet1')

li = (list(sheet.columns)[column_index_from_string('AA')])[4:] #前4行丢掉
row_n = len(li) #查看总共有多少行
result_row = 1

sheet_result.cell(row=result_row,column=1).value = '土石方开挖工程量（m³）'
sheet_result.cell(row=result_row,column=2).value = '回填深度（m）'
sheet_result.cell(row=result_row,column=3).value = '管沟砂垫层（m³）'
sheet_result.cell(row=result_row,column=4).value = ' 砂砾石回填（m³）'
sheet_result.cell(row=result_row,column=5).value = ' 土方回填深度（m）'
sheet_result.cell(row=result_row,column=6).value = '土方回填体积（m³）'
sheet_result.cell(row=result_row,column=7).value = '沟槽顶宽（m）'
sheet_result.cell(row=result_row,column=8).value = '拆除及恢复路面面积（㎡）'
sheet_result.cell(row=result_row,column=column_index_from_string('M')).value = '建渣（m³）'
sheet_result.cell(row=result_row,column=14).value = '余土外运体积（m³）'
sheet_result.cell(row=result_row,column=17).value='土石方开挖深度（m）Z'
for i in range(1,row_n,2):
    r = 4+i
    result_row += 1

    X = F(sheet['X'+str(r)].value)
    Y = F(sheet['Y'+str(r)].value)
    AL = F(sheet['AL'+str(r)].value)
    AK = F(sheet['AK'+str(r)].value)
    AJ = F(sheet['AJ'+str(r)].value)
    W = F(sheet['W' + str(r)].value)
    Z = W-AK-AJ-AL
    S = F(sheet['S'+str(r)].value)
    T = F(sheet['T'+str(r)].value)
    U = F(sheet['U'+str(r)].value)
    AG = F(sheet['AG'+str(r)].value)
    Q = F(sheet['Q'+str(r)].value)

    #Z
    sheet_result.cell(row=result_row, column=17).value = Z

    #AA
    if sheet.cell(row=r,column=column_index_from_string('Y')).value == '/':
        sheet_result.cell(row=result_row,column=1).value =AA = (X+AG)*Z/2*(T+U)
    elif sheet.cell(row=r,column=column_index_from_string('X')).value == '基坑开挖':
        sheet_result.cell(row=result_row, column=1).value = AA = (2.8+Y*W)*(2.1+Y*W)*W+1/3*Y*Y*W*W*W
    else:
        sheet_result.cell(row=result_row, column=1).value = AA = (X+Y*Z)*Z*(S+T+U)
    #AB
    sheet_result.cell(row=result_row,column=2).value = AB = Z

    #AC
    if sheet.cell(row=r, column=column_index_from_string('Y')).value == '/':
        sheet_result.cell(row=result_row,column=3).value = AC = X*(0.15+1/4*Q)*T
    else:
        sheet_result.cell(row=result_row,column=3).value = AC = (X+Y*(0.15+1/4*Q))*(0.15+1/4*Q)*(S+T)
    #AE
    if sheet.cell(row=r,column=column_index_from_string('AK')).value != None:
        AE = AB - 0.95 -Q
    else:
        AE = AB - 0.65 - Q
    if AE <= 0 or AE < 0.0008:
        AE = 0
    sheet_result.cell(row=result_row,column=5).value = AE
    #AD
    if sheet.cell(row=r,column=column_index_from_string('Y')).value == '/':
        sheet_result.cell(row=result_row,column=4).value= AD =(X+AG)*(AB-AE)/2*T-AC
    else:
        sheet_result.cell(row=result_row,column=4).value= AD =(X+Y*(AB-AE))*(AB-AE)*(S+T)-AC

    # AF  AN
    if sheet.cell(row=r,column=column_index_from_string('Y')).value == '/':
        sheet_result.cell(row=result_row,column=6).value = AF = (X+AG)*AB/2*T-AC-AD
        sheet_result.cell(row=result_row,column=14).value = AN = AA - AF
    elif get_source_value(r,'AC') == None and get_source_value(r,'AE') == None and get_source_value(r,'AD') == None and get_source_value(r,'AB')==None:
        sheet_result.cell(row=result_row,column=6).value=None
        AF = 0
        sheet_result.cell(row=result_row,column=14).value=AA-AF
    else:
        if get_source_value(r,'AC') == None and get_source_value(r,'AE') == None and get_source_value(r,'AD') == None:
                AN = ((0.5+Q)*0.2+(0.3+Q)*(0.3+Q))*U
                sheet_result.cell(row=result_row,column=6).value = AF = AA-AN
                sheet_result.cell(row=result_row,column=14).value = AN
        else:
            sheet_result.cell(row=result_row,column=6).value = AF = (X+Y*AB)*AB*(S+T)-AC-AD
            AN = AA-AF
            sheet_result.cell(row=result_row,column=14).value = AN

    # AH
    if get_source_value(r,'AG') != None:
        if sheet.cell(row=r,column=column_index_from_string('Y')) == '/':
            sheet_result.cell(row=result_row,column=8).value = AH = AG*T
        else:
            sheet_result.cell(row=result_row,column=8).value = AH = AG*(S+T)
    #AM
    if get_source_value(r,'AG') != None:
        sheet_result.cell(row=result_row,column=column_index_from_string('M')).value = AH*(AJ+AK+AL)


wb_result.save('result1.xlsx')
print('End....')
